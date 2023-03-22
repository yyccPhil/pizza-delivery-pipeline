from openpyxl import load_workbook
import csv

with open("state_population.csv", 'w') as file:
    writer = csv.writer(file)

    ws = load_workbook("NST-EST2022-POP.xlsx").active
    for row in range(10, 61):
        writer.writerow([
            ws[chr(65) + str(row)].value.strip().replace('.', ''),
            ws[chr(69) + str(row)].value
        ])
